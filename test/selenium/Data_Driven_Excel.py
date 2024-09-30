from time import sleep

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

# Load the Excel Sheet
workbook = load_workbook("Python_XL_workbook.xlsx")
# Selecting Active sheet
sheet = workbook.active

driver = webdriver.Firefox()
driver.maximize_window()
# Iterating the values like Username and Password
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]
    driver.get("https://www.saucedemo.com/v1/")
    sleep(3)
    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    sleep(3)
    driver.find_element(By.ID,"login-button").click()
    sleep(3)
    driver.find_element(By.XPATH,"//button[normalize-space()='Open Menu']").click()
    sleep(3)
    driver.find_element(By.XPATH, "//a[@id='logout_sidebar_link']").click()
    sleep(3)

driver.quit()





